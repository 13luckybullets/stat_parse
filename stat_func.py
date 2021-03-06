#coding:utf-8

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles.borders import Border, Side


SHEET_DATA = {}
COLUMN_NAMES = ['H_S', 'H_M', 'A_S', 'A_M']
DA_INDEX = column_index_from_string('DA')


def check_headers(sheet):
    headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column+1)]
    if 'H_S1' in headers:
        return True
    return False


def prepare_sheet(sheet):
    column = DA_INDEX

    for name in COLUMN_NAMES:
        for i in range(1, 8, 1):
            sheet.cell(row=1, column=column).value = "%s%s" % (name, i)
            column += 1


def get_by_teams(sheet, row_num, name):
    data = {'score': [], 'missed': []}

    for j in range(row_num, 1, -1):
        if name == sheet.cell(row=j, column=3).value:
            data['score'].append(sheet.cell(row=j, column=5).value)
            data['missed'].append(sheet.cell(row=j, column=6).value)
        elif name == sheet.cell(row=j, column=4).value:
            data['score'].append(sheet.cell(row=j, column=6).value)
            data['missed'].append(sheet.cell(row=j, column=5).value)
        elif len(data['score']) == 7:
            break
    return data


def update_sheet_data(row_num, team, data):
    if data and data['score']:
        if len(data['score']) >= 7 or len(data['missed']) >= 7:
            data['score'] = data['score'][:7]
            data['missed'] = data['missed'][:7]

        SHEET_DATA.update({
            tuple((row_num, team)): data
        })


def take_data(sheet, row_num):
    name_home = sheet.cell(row=row_num, column=3).value
    name_guest = sheet.cell(row=row_num, column=4).value

    data_home = get_by_teams(sheet, row_num-1, name_home)
    data_guest = get_by_teams(sheet, row_num-1, name_guest)

    update_sheet_data(row_num, 3, data_home)
    update_sheet_data(row_num, 4, data_guest)


def set_border(sheet, row_num, column):
    sheet.cell(row=row_num, column=column).border = Border(
        left=Side(style='medium'))


def write_data(sheet):
    for val in SHEET_DATA:
        data = SHEET_DATA.get(val)
        team = val[1]
        row_num = val[0]

        if team == 3:
            column_score = DA_INDEX
        else:
            column_score = DA_INDEX + 14

        column_miss = column_score + 7

        set_border(sheet, row_num, column_score)
        for d in data.get('score'):
            sheet.cell(row=row_num, column=column_score).value = d
            column_score += 1

        set_border(sheet, row_num, column_miss)
        for d in data.get('missed'):
            sheet.cell(row=row_num, column=column_miss).value = d
            column_miss += 1
